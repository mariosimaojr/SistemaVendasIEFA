from __future__ import annotations

import argparse
import re
import unicodedata
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_INPUT = Path(r"C:\Users\mario\Downloads\CEFA inventario.xlsx")
DEFAULT_INPUT_ACCENTED = Path(r"C:\Users\mario\Downloads\CEFA inventário.xlsx")
DEFAULT_OUTPUT = Path(__file__).resolve().parent / "03 - Carga inventario CEFA.sql"

CATEGORY_ORDER = [
    "ÁGUA",
    "ALIMENTO",
    "BEBÊ",
    "BIJUTERIA",
    "BOLSA",
    "BRINQUEDO",
    "COZINHA",
    "DECORAÇÃO",
    "DIVERSOS",
    "ENXOVAL",
    "HIGIENE",
    "LIVRO",
    "LIVRO INFANTIL",
    "PET",
    "RELIGIOSO",
    "VESTUÁRIO",
]


@dataclass
class ProductRow:
    source_sheet: str
    source_row: int
    category: str
    name: str
    description: str
    price: Decimal
    quantity: int
    missing_price: bool


@dataclass
class SkippedRow:
    source_sheet: str
    source_row: int
    reason: str


def clean_text(value: Any) -> str:
    if value is None:
        return ""

    return re.sub(r"\s+", " ", str(value).strip())


def normalize_key(value: Any) -> str:
    text = clean_text(value).lower()
    text = "".join(
        char
        for char in unicodedata.normalize("NFD", text)
        if unicodedata.category(char) != "Mn"
    )

    return text


def normalize_category(value: Any) -> str:
    return clean_text(value).upper()


def value_by_header(row: dict[str, Any], *headers: str) -> Any:
    normalized_row = {
        normalize_key(header): value
        for header, value in row.items()
        if clean_text(header)
    }

    for header in headers:
        value = normalized_row.get(normalize_key(header))
        if value is not None:
            return value

    return None


def decimal_or_zero(value: Any) -> tuple[Decimal, bool]:
    if value is None or clean_text(value) == "":
        return Decimal("0.00"), True

    text = clean_text(value).replace(",", ".")

    try:
        decimal_value = Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"preco invalido: {value!r}") from exc

    return decimal_value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP), False


def quantity_or_zero(value: Any) -> int:
    if value is None or clean_text(value) == "":
        return 0

    if isinstance(value, float) and not value.is_integer():
        raise ValueError(f"quantidade nao inteira: {value!r}")

    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"quantidade invalida: {value!r}") from exc


def join_name(*parts: Any) -> str:
    return " - ".join(clean_text(part) for part in parts if clean_text(part))


def title_case_name(value: str) -> str:
    text = clean_text(value).lower()

    return re.sub(
        r"(^|[\s/\-:('\"])([a-zà-öø-ÿ])",
        lambda match: match.group(1) + match.group(2).upper(),
        text,
    )


def source_line(sheet_name: str, row_number: int) -> str:
    return f"Origem: aba {sheet_name}, linha {row_number}"


def build_description(
    sheet_name: str,
    row_number: int,
    row: dict[str, Any],
    missing_price: bool,
) -> str:
    details: list[str] = [source_line(sheet_name, row_number)]

    author = clean_text(value_by_header(row, "Autor"))
    if author:
        details.append(f"Autor: {author}")

    for label, header in [
        ("Produto", "produto"),
        ("Titulo", "Titulo"),
        ("Formato", "formato"),
        ("Medida", "medida"),
        ("Medida", "medida (metro)"),
        ("Sabor", "sabor"),
        ("Obs", "obs"),
    ]:
        value = clean_text(value_by_header(row, header))
        if value:
            details.append(f"{label}: {value}")

    if missing_price:
        details.append("Preco ausente na planilha; carregado como 0.00.")

    return "; ".join(dict.fromkeys(details))


def build_product_name(sheet_name: str, row: dict[str, Any]) -> str:
    if sheet_name in {"livros", "livros infantis"}:
        return title_case_name(clean_text(value_by_header(row, "Titulo")))

    if sheet_name == "bazar":
        return title_case_name(
            join_name(
                value_by_header(row, "produto"),
                value_by_header(row, "formato"),
                value_by_header(row, "medida (metro)"),
                value_by_header(row, "obs"),
            )
        )

    if sheet_name == "alimentos":
        return title_case_name(
            join_name(
                value_by_header(row, "produto"),
                value_by_header(row, "sabor"),
                value_by_header(row, "obs"),
            )
        )

    if sheet_name == "vestuário":
        return title_case_name(
            join_name(
                value_by_header(row, "produto"),
                value_by_header(row, "medida"),
                value_by_header(row, "obs"),
            )
        )

    return title_case_name(clean_text(value_by_header(row, "produto", "Titulo")))


def iter_inventory_rows(workbook_path: Path) -> tuple[list[ProductRow], list[SkippedRow]]:
    workbook = load_workbook(workbook_path, data_only=True)
    products: list[ProductRow] = []
    skipped_rows: list[SkippedRow] = []

    for sheet in workbook.worksheets:
        headers = [clean_text(cell.value) for cell in sheet[1]]

        for row_number in range(2, sheet.max_row + 1):
            values = [
                sheet.cell(row_number, column_number).value
                for column_number in range(1, sheet.max_column + 1)
            ]

            if not any(clean_text(value) for value in values):
                continue

            row = dict(zip(headers, values))
            category = normalize_category(value_by_header(row, "categoria"))
            name = build_product_name(sheet.title, row)

            if not category and not name:
                skipped_rows.append(
                    SkippedRow(sheet.title, row_number, "linha sem categoria e produto")
                )
                continue

            if not category or not name:
                skipped_rows.append(SkippedRow(sheet.title, row_number, "linha incompleta"))
                continue

            price, missing_price = decimal_or_zero(
                value_by_header(row, "preco", "preço", "valor unitario", "valor unitário")
            )
            quantity = max(0, quantity_or_zero(value_by_header(row, "qde", "qde estoque")))

            products.append(
                ProductRow(
                    source_sheet=sheet.title,
                    source_row=row_number,
                    category=category,
                    name=name,
                    description=build_description(
                        sheet.title,
                        row_number,
                        row,
                        missing_price,
                    ),
                    price=price,
                    quantity=quantity,
                    missing_price=missing_price,
                )
            )

    return products, skipped_rows


def sql_string(value: str) -> str:
    escaped = (
        value.replace("\\", "\\\\")
        .replace("'", "''")
        .replace("\r\n", "\\n")
        .replace("\n", "\\n")
        .replace("\r", "\\n")
    )

    return f"'{escaped}'"


def sql_decimal(value: Decimal) -> str:
    return f"{value:.2f}"


def ordered_categories(products: list[ProductRow]) -> list[str]:
    categories = {product.category for product in products}
    ordered = [category for category in CATEGORY_ORDER if category in categories]
    remaining = sorted(categories - set(ordered), key=normalize_key)

    return ordered + remaining


def comma_values(rows: list[str]) -> str:
    return ",\n".join(rows)


def render_sql(products: list[ProductRow], skipped_rows: list[SkippedRow]) -> str:
    categories = ordered_categories(products)
    category_ids = {category: index for index, category in enumerate(categories, start=1)}

    category_values = [
        f"({category_id}, {sql_string(category)}, {sql_string(category)}, 1)"
        for category, category_id in category_ids.items()
    ]

    product_values = [
        (
            f"({product_id}, {sql_string(product.name)}, "
            f"{sql_string(product.description)}, {category_ids[product.category]}, "
            f"{sql_decimal(product.price)}, 1, NOW())"
        )
        for product_id, product in enumerate(products, start=1)
    ]

    movement_values = [
        (
            f"({movement_id}, {product_id}, {product.quantity}, NOW(), "
            f"{sql_string(f'Carga inicial inventario CEFA - aba {product.source_sheet}, linha {product.source_row}')}, "
            "1, 'ENTRADA')"
        )
        for movement_id, (product_id, product) in enumerate(
            (
                (product_id, product)
                for product_id, product in enumerate(products, start=1)
                if product.quantity > 0
            ),
            start=1,
        )
    ]

    total_quantity = sum(product.quantity for product in products)
    missing_price_count = sum(1 for product in products if product.missing_price)
    without_movement_count = sum(1 for product in products if product.quantity <= 0)

    lines = [
        "-- Carga do inventario CEFA",
        "-- Gerado por BD/gerar_carga_inventario_cefa.py",
        "-- Fonte: CEFA inventario.xlsx",
        f"-- Categorias: {len(categories)}",
        f"-- Produtos: {len(products)}",
        f"-- Movimentacoes de entrada: {len(movement_values)}",
        f"-- Unidades em estoque: {total_quantity}",
        f"-- Produtos com preco ausente carregados como 0.00: {missing_price_count}",
        f"-- Produtos sem movimentacao de estoque: {without_movement_count}",
        f"-- Linhas ignoradas: {len(skipped_rows)}",
        "",
        "USE `controle_vendas_iefa`;",
        "SET NAMES utf8mb4;",
        "",
        "START TRANSACTION;",
        "",
        "DELETE FROM `venda_itens`;",
        "DELETE FROM `vendas`;",
        "DELETE FROM `movimentacoes_estoque`;",
        "DELETE FROM `produtos`;",
        "DELETE FROM `categorias`;",
        "",
        "INSERT INTO `categorias` (`SEQUENCIA`, `NOME`, `DESCRICAO`, `ATIVO`) VALUES",
        comma_values(category_values) + ";",
        "",
        (
            "INSERT INTO `produtos` "
            "(`SEQUENCIA`, `NOME`, `DESCRICAO`, `SEQCATEGORIA`, "
            "`PRECO_VENDA`, `ATIVO`, `DATA_CADASTRO`) VALUES"
        ),
        comma_values(product_values) + ";",
    ]

    if movement_values:
        lines.extend(
            [
                "",
                (
                    "INSERT INTO `movimentacoes_estoque` "
                    "(`SEQUENCIA`, `SEQPRODUTO`, `QUANTIDADE`, `DATA_MOVIMENTO`, "
                    "`OBSERVACAO`, `SEQUSUARIO`, `TIPO_MOVIMENTO`) VALUES"
                ),
                comma_values(movement_values) + ";",
            ]
        )

    lines.extend(
        [
            "",
            "COMMIT;",
            "",
            f"ALTER TABLE `categorias` AUTO_INCREMENT = {len(categories) + 1};",
            f"ALTER TABLE `produtos` AUTO_INCREMENT = {len(products) + 1};",
            f"ALTER TABLE `movimentacoes_estoque` AUTO_INCREMENT = {len(movement_values) + 1};",
            "ALTER TABLE `venda_itens` AUTO_INCREMENT = 1;",
            "ALTER TABLE `vendas` AUTO_INCREMENT = 1;",
            "",
            "-- Conferencias esperadas apos a carga:",
            "SELECT COUNT(*) AS categorias FROM `categorias`;",
            "SELECT COUNT(*) AS produtos FROM `produtos`;",
            "SELECT COUNT(*) AS movimentacoes_estoque FROM `movimentacoes_estoque`;",
            "SELECT COALESCE(SUM(`QUANTIDADE`), 0) AS estoque_total FROM `movimentacoes_estoque`;",
            "SELECT COUNT(*) AS vendas FROM `vendas`;",
            "SELECT COUNT(*) AS venda_itens FROM `venda_itens`;",
            "",
        ]
    )

    return "\n".join(lines)


def print_summary(products: list[ProductRow], skipped_rows: list[SkippedRow]) -> None:
    categories = ordered_categories(products)
    movements = [product for product in products if product.quantity > 0]
    total_quantity = sum(product.quantity for product in products)
    missing_price_count = sum(1 for product in products if product.missing_price)
    without_movement_count = sum(1 for product in products if product.quantity <= 0)

    print(f"Categorias: {len(categories)}")
    print(f"Produtos: {len(products)}")
    print(f"Movimentacoes de entrada: {len(movements)}")
    print(f"Unidades em estoque: {total_quantity}")
    print(f"Produtos com preco ausente carregados como 0.00: {missing_price_count}")
    print(f"Produtos sem movimentacao de estoque: {without_movement_count}")
    print(f"Linhas ignoradas: {len(skipped_rows)}")

    if skipped_rows:
        print("Linhas ignoradas detalhadas:")
        for skipped_row in skipped_rows:
            print(
                f"- {skipped_row.source_sheet} linha {skipped_row.source_row}: "
                f"{skipped_row.reason}"
            )


def validate_summary(products: list[ProductRow], skipped_rows: list[SkippedRow]) -> None:
    expected = {
        "categorias": 16,
        "produtos": 500,
        "movimentacoes": 483,
        "unidades": 1643,
        "linhas_ignoradas": 4,
    }
    actual = {
        "categorias": len(ordered_categories(products)),
        "produtos": len(products),
        "movimentacoes": sum(1 for product in products if product.quantity > 0),
        "unidades": sum(product.quantity for product in products),
        "linhas_ignoradas": len(skipped_rows),
    }

    mismatches = [
        f"{key}: esperado {expected[key]}, obtido {actual[key]}"
        for key in expected
        if expected[key] != actual[key]
    ]

    if mismatches:
        raise RuntimeError("Resumo fora do esperado: " + "; ".join(mismatches))


def parse_args() -> argparse.Namespace:
    default_input = DEFAULT_INPUT_ACCENTED if DEFAULT_INPUT_ACCENTED.exists() else DEFAULT_INPUT

    parser = argparse.ArgumentParser(
        description="Gera o SQL de carga do inventario CEFA a partir da planilha."
    )
    parser.add_argument(
        "--input",
        default=str(default_input),
        help="Caminho da planilha CEFA inventario.xlsx.",
    )
    parser.add_argument(
        "--output",
        default=str(DEFAULT_OUTPUT),
        help="Caminho do arquivo SQL de saida.",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)

    products, skipped_rows = iter_inventory_rows(input_path)
    validate_summary(products, skipped_rows)

    sql = render_sql(products, skipped_rows)
    output_path.write_text(sql, encoding="utf-8", newline="\n")

    print_summary(products, skipped_rows)
    print(f"SQL gerado em: {output_path}")


if __name__ == "__main__":
    main()
